#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import  PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side

class NavXlsxFile:
    """ navigation xlsx file """
    def __init__(self):
        # self.cfg_file = cfg_file
        self.work_book = Workbook()
        self.current_sheet = self.work_book.active
        self.current_sheet.title = 'navLog'
        self.cursor = {'row':1, 'col':1}
        self.sheet_style =  NamedStyle(name='sheet_style')
        self.sheet_style.alignment = Alignment(horizontal='center', vertical = 'center')
        bd = Side(style='thin', color="000000")
        # self.sheet_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.work_book.add_named_style(self.sheet_style)


    def createSheet(self, sheet_name):
        work_sheet = self.work_book.create_sheet(sheet_name, 0)
        self.selectSheet(sheet_name)
        return work_sheet
    

    def selectSheet(self, sheet_name):
        self.current_sheet = self.work_book[sheet_name]
        return self.current_sheet

    
    def setCursor(self, row, col):
        if row >= 1 and col >= 1:
            self.cursor['row'] = row
            self.cursor['col'] = col


    def writeCell(self, row, col, value):
        work_sheet = self.current_sheet
        work_sheet.cell(row, col, value).style = self.sheet_style


    def resize(self, sheet_name):
        """ resize sheet unit automotive """
        try:
            for col in range(self.current_sheet.max_column):
                col_width = 0
                for row in range(self.current_sheet.max_row):
                    cell_value = self.current_sheet.cell(row+1, col+1).value
                    if len(str(cell_value)) > col_width:
                        col_width = len(str(cell_value))

                self.current_sheet.column_dimensions[get_column_letter(col+1)].width\
                    = float(col_width + 3.0)
        except TypeError:
            print(self.current_sheet.max_column)


    def create(self, file_path):
        """ create .xlsx file in file system """
        try:
            self.work_book.save(file_path)
        except:
            print('create xlsx file faild!')
            sys.exit()


    def setCellColor(self, sheet_name, row, col, color):
        ws = self.work_book[sheet_name]
        cell = ws.cell(row, col)
        fill = PatternFill("solid", fgColor=color)
        cell.fill = fill


    def setCellBorder(self, sheet_name, begin_row, begin_col, end_row, end_col):
        ws = self.work_book[sheet_name]
        cells = [(cell_row, cell_col) for cell_row in range(begin_row, end_row + 1)
                                      for cell_col in range(begin_col, end_col + 1)]
        for cur_cell in cells:
            ws.cell(cur_cell[0], cur_cell[1]).border = self.border
        

    def mergeCell(self, sheet_name, begin_row, end_row, begin_col=1, end_col=3):
        ws = self.work_book[sheet_name]
        cells = [(cell_row, cell_col) for cell_col in range(begin_col, end_col+1)
                                      for cell_row in range(begin_row, end_row+1)]
        
        for cur_col in range(begin_col, end_col+1):
            cells = [(cell_row, cur_col) for cell_row in range(begin_row, end_row+1)]

            merge_flag = 0
            blank_index = 0
            cell_index = 0

            while cell_index < len(cells):
                cur_cell = cells[cell_index]
                if ws.cell(cur_cell[0], cur_cell[1]).value != None:
                    pass
                    if merge_flag == 0:
                        merge_flag = 1
                    else:
                        merge_flag = 0
                        if blank_index > 0:
                            ws.merge_cells(start_row=cur_cell[0]-(blank_index+1), start_column=cur_cell[1],
                                        end_row=cur_cell[0]-1, end_column=cur_cell[1])
                            cell_index -= 1
                            blank_index = 0
                        else:
                            cell_index -= 1
                else:
                    pass
                    blank_index += 1
                    if cell_index == len(cells) - 1:
                        if merge_flag > 0:
                            ws.merge_cells(start_row=cur_cell[0]-blank_index, start_column=cur_cell[1],
                                           end_row=cur_cell[0], end_column=cur_cell[1])
                cell_index += 1



