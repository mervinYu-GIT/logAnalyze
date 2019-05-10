#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
from navLog import NavLogFile


class NavXlsxFile:
    """ navigation xlsx file """
    def __init__(self):
        # self.cfg_file = cfg_file
        self.work_book = Workbook()
        self.current_sheet = self.work_book.active
        self.current_sheet.title = 'navLog'
        self.cursor = {'row':1, 'col':1}
        self.sheet_style =  NamedStyle(name='sheet_style')
        self.sheet_style.alignment = Alignment(horizontal='center', vertical = 'center')
        bd = Side(style='thin', color="000000")
        self.sheet_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.work_book.add_named_style(self.sheet_style)


    def create_sheet(self, sheet_name):
        return self.work_book.create_sheet(sheet_name, 0)
    

    def select_sheet(self, sheet_name):
        self.current_sheet = self.work_book[sheet_name]
        return self.current_sheet

    
    def set_cursor(self, row, col):
        if row >= 1 and col >= 1:
            self.cursor['row'] = row
            self.cursor['col'] = col


    def write_cell(self, row, col, value):
        work_sheet = self.current_sheet
        work_sheet.cell(row, col, value).style = self.sheet_style


    def resize(self, sheet_name):
        """ resize sheet unit automotive """
        for col in range(self.current_sheet.max_column):
            col_width = 0
            for row in range(self.current_sheet.max_row):
                cell_value = self.current_sheet.cell(row+1, col+1).value
                if len(str(cell_value)) > col_width:
                    col_width = len(str(cell_value))

            self.current_sheet.column_dimensions[get_column_letter(col+1)].width\
                = float(col_width + 3.0)


    def create(self, file_path):
        """ create .xlsx file in file system """
        try:
            self.work_book.save(file_path)
        except:
            print('create xlsx file faild!')
            sys.exit()


    def multiple_rounds(self, log_file, origin_cursor, start, end):
        nav_log_file = NavLogFile(log_file)
        searchs = nav_log_file.searchLogs("Message", start, end)
        index = 2
        round_index = 1
        row = origin_cursor['row']
        col = origin_cursor['col'] + 4
        offset = 0
        while index < len(searchs) - 1:
            round_index += 1
            self.write_cell(row, col + offset, 'time cost(s)Round' + str(round_index))
            begin = datetime.strptime(searchs[index][nav_log_file.logHeadRow.index('Time')],\
                "%d.%m.%Y %H:%M:%S:%f")
            end = datetime.strptime(searchs[index+1][nav_log_file.logHeadRow.index('Time')],\
                "%d.%m.%Y %H:%M:%S:%f")
            delta_time = end - begin
            self.write_cell(row + 1, col + offset, delta_time.total_seconds())
            offset += 1
            index += 2

