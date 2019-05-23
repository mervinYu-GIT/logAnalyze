#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import  PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
from navLog import NavLogFile


class NavXlsxFile:
    """ navigation xlsx file """
    def __init__(self):
        # self.cfg_file = cfg_file
        self.work_book = Workbook()
        self.current_sheet = self.work_book.active
        self.current_sheet.title = 'navLog'
        self.cursor = {'row':1, 'col':1}
        self.sheet_style =  NamedStyle(name='sheet_style')
        self.sheet_style.alignment = Alignment(horizontal='center', vertical = 'center')
        bd = Side(style='thin', color="000000")
        self.sheet_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.work_book.add_named_style(self.sheet_style)


    def createSheet(self, sheet_name):
        work_sheet = self.work_book.create_sheet(sheet_name, 0)
        self.selectSheet(sheet_name)
        return work_sheet
    

    def selectSheet(self, sheet_name):
        self.current_sheet = self.work_book[sheet_name]
        return self.current_sheet

    
    def setCursor(self, row, col):
        if row >= 1 and col >= 1:
            self.cursor['row'] = row
            self.cursor['col'] = col


    def writeCell(self, row, col, value):
        work_sheet = self.current_sheet
        work_sheet.cell(row, col, value).style = self.sheet_style


    def resize(self, sheet_name):
        """ resize sheet unit automotive """
        for col in range(self.current_sheet.max_column):
            col_width = 0
            for row in range(self.current_sheet.max_row):
                cell_value = self.current_sheet.cell(row+1, col+1).value
                if len(str(cell_value)) > col_width:
                    col_width = len(str(cell_value))

            self.current_sheet.column_dimensions[get_column_letter(col+1)].width\
                = float(col_width + 3.0)


    def create(self, file_path):
        """ create .xlsx file in file system """
        try:
            self.work_book.save(file_path)
        except:
            print('create xlsx file faild!')
            sys.exit()


    def setCellColor(self, sheet_name, row, col, color):
        ws = self.work_book[sheet_name]
        cell = ws.cell(row, col)
        fill = PatternFill("solid", fgColor=color)
        cell.fill = fill



