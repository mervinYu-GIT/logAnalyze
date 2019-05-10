#!/usr/bin/python
# -*- coding: utf-8 -*-

import os, sys
from os import path
import json
from datetime import datetime
from argparse import ArgumentParser
from navXlsx import NavXlsxFile
from navLog import NavLogFile
from openpyxl import Workbook
import public_fun
import collections


class NavLogAnalyze:
    pass



if __name__ == "__main__":
    arg_parser = ArgumentParser()
    arg_parser.add_argument('log_files', nargs = '*', help = "input log file")
    arg_parser.add_argument('--cfg', help="config file, eg: path/config.json")
    arg_parser.add_argument('--xlsx', help='xlsx file')
    args = arg_parser.parse_args()

    if args.log_files:
        log_files = args.log_files
    if args.cfg:
        cfg_file = args.cfg
    else:
        cfg_file = './config.json'

    if args.xlsx:
        xlsx_file_path = args.xlsx
    else:
        xlsx_file_path = './navLog.xlsx'

    xlsx_file = NavXlsxFile(xlsx_file_path, cfg_file)
    with open(cfg_file, 'r') as json_file:
        try:
            json_data = json.load(json_file, object_pairs_hook=collections.OrderedDict)
        except:
            print(cfg_file + ' is not json format!')
            sys.exit()

        for log_file in log_files:
            nav_log_file = NavLogFile(log_file)
            sheet_name = path.basename(log_file).split('.')[0]
            xlsx_file.create_sheet(sheet_name)
            work_sheet = xlsx_file.select_sheet(sheet_name)
            
            origin_point = {'row':1, 'col':1}
            for k, v in json_data.items():
                row_offset = 0
                col_offset = 0
                route_flag = 0
                search_flag = 0
               
                if k == 'Routing':
                    route_flag = 1       
                    route_cursor = {'row':origin_point['row'] + row_offset,
                                    'col':origin_point['col'] + col_offset}        
                
                if k == 'Search':
                    search_flag = 1
                    search_cursor = {'row':origin_point['row'] + row_offset,
                                     'col': origin_point['col'] + col_offset}
                
                xlsx_file.write_cell(origin_point['row'] ,origin_point['col'], k)
                col_offset += 1
                row_offset += 1
                index = 1

                for k_1, v_1 in v.items():
                    xlsx_file.write_cell(origin_point['row'] + row_offset,\
                        origin_point['col'] - 1 + col_offset, index)
                    index += 1
                    for k_2, v_2 in v_1.items():
                        if k_2 == 'log point':
                            if work_sheet.cell(origin_point['row'], origin_point['col'] + col_offset).value == None:
                                xlsx_file.write_cell(origin_point['row'],\
                                    origin_point['col'] + col_offset, "time cost(s)Round1") 

                            if v_2['begin'] == '':        # search begin time in file logs
                                beginTime = nav_log_file.beginTime()
                            else:
                                beginTime = nav_log_file.searchTime(v_2['begin'].encode('utf-8'), 'Message')

                            if v_2['end'] == '':          # search end time
                                endTime = nav_log_file.endTime()
                            else:
                                endTime = nav_log_file.searchTime(v_2['end'].encode('utf-8'), 'Message')

                            if beginTime == -1 or endTime == -1:   # calc delta time
                                col_offset += 1
                                continue
                            else:
                                delta_time = public_fun.calcTime(beginTime, endTime)
                                xlsx_file.write_cell(origin_point['row'] + row_offset,\
                                    origin_point['col'] + col_offset, delta_time.total_seconds())
                                col_offset += 1

                        else:   # <if k_2 == 'log point'>
                            xlsx_file.write_cell(origin_point['row'] + row_offset,\
                                origin_point['col'] + col_offset, v_2)
                                    
                            if work_sheet.cell(origin_point['row'], origin_point['col'] + col_offset).value == None:
                                xlsx_file.write_cell(origin_point['row'],\
                                    origin_point['col'] + col_offset, k_2)  
                            col_offset += 1
                    row_offset += 1
                    col_offset -= len(v_1)
                origin_point['row'] = work_sheet.max_row + 3

                if route_flag == 1:
                    searchs = nav_log_file.searchLogs("Message", \
                        json_data['Search']['OneBox']['log point']['begin'].encode('utf-8'),\
                        json_data['Search']['OneBox']['log point']['end'].encode('utf-8'))
                    index = 2
                    round_index = 1
                    row = route_cursor['row']
                    col = route_cursor['col'] + 4
                    offset = 0
                    while index < len(searchs) - 1:
                        round_index += 1
                        xlsx_file.write_cell(row, col + offset, 'time cost(s)Round' + str(round_index))
                        begin = datetime.strptime(searchs[index][nav_log_file.logHeadRow['Time'].index()],\
                            "%d.%m.%Y %H:%M:%S:%f")
                        end = datetime.strptime(searchs[index+1][nav_log_file.logHeadRow['Time'].index()],\
                            "%d.%m.%Y %H:%M:%S:%f")
                        delta_time = end - begin
                        xlsx_file.write_cell(row + 1, col + offset, delta_time.total_seconds())
                        offset += 1
 
                if search_flag == 1:
                    searchs = nav_log_file.searchLogs("Message", \
                        json_data['Search']['OneBox']['log point']['begin'].encode('utf-8'),\
                        json_data['Search']['OneBox']['log point']['end'].encode('utf-8'))
                    index = 2
                    round_index = 1
                    row = search_cursor['row']
                    col = search_cursor['col'] + 4
                    offset = 0
                    while index < len(searchs) - 1:
                        round_index += 1
                        xlsx_file.write_cell(row, col + offset, 'time cost(s)Round' + str(round_index))
                        begin = datetime.strptime(searchs[index][nav_log_file.logHeadRow['Time'].index()],\
                            "%d.%m.%Y %H:%M:%S:%f")
                        end = datetime.strptime(searchs[index+1][nav_log_file.logHeadRow['Time'].index()],\
                            "%d.%m.%Y %H:%M:%S:%f")
                        delta_time = end - begin
                        xlsx_file.write_cell(row + 1, col + offset, delta_time.total_seconds())
                        offset += 1

    xlsx_file.resize(sheet_name)
    xlsx_file.create(xlsx_file_path)


